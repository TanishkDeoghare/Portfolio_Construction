import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load your data into a pandas DataFrame
file_path = r"C:/Users/tanishk.deoghare/OneDrive - Angel Oak Capital Advisors/Desktop/Portfolio Construction/Dry Run 1/Merged_Spreads_Report.xlsx"
df = pd.read_excel(file_path)

date_col = 'Date'

# Check if the date column exists
if date_col not in df.columns:
    raise KeyError(f"Date column '{date_col}' not found in DataFrame")

# Get the length of the date column
date_col_length = len(df[date_col])

# Define a function to interpolate or extrapolate based on trendline
def fill_missing_with_trend(df, y_col, x_col):
    # Extract the columns
    y = df[y_col]
    x = df[x_col]

    # Find the non-missing indices
    non_missing_indices = y.dropna().index

    # Perform linear regression to find the slope and intercept
    slope, intercept = np.polyfit(x[non_missing_indices], y.dropna(), 1)

    # Define a function for the trendline
    trendline = lambda x_val: slope * x_val + intercept

    # Fill in missing values using the trendline
    y_filled = y.copy()
    y_filled[y.isna()] = x[y.isna()].apply(trendline)

    return y_filled

# Find a suitable column for the x-axis (default to the first numerical column if 'Date' is not numerical)
default_x_col = date_col if np.issubdtype(df[date_col].dtype, np.number) else df.select_dtypes(include=np.number).columns[0]

# Find columns with length less than the length of the date column and apply the trendline function
for column in df.columns:
    if df[column].count() < date_col_length and column != date_col and column != default_x_col:
        df[column] = fill_missing_with_trend(df, column, default_x_col)

# Save the updated DataFrame to a new Excel file and highlight the filled cells
output_path = r"C:/Users/tanishk.deoghare/OneDrive - Angel Oak Capital Advisors/Desktop/Portfolio Construction/Dry Run 1/Filled_Merged_Spreads_Report.xlsx"
df.to_excel(output_path, index=False)

# Load the workbook and worksheet
wb = load_workbook(output_path)
ws = wb.active

# Define the fill for highlighting
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Highlight the filled cells
for col in df.columns:
    for row in range(2, len(df) + 2):  # Assuming the first row is the header
        cell = ws.cell(row=row, column=df.columns.get_loc(col) + 1)
        if pd.isna(df.at[row-2, col]):
            cell.fill = fill

# Save the workbook with highlighted cells
wb.save(output_path)

print(f"Missing values filled and highlighted. The updated file is saved at {output_path}")
